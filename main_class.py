import os
import shutil
import psutil
import openpyxl
from openpyxl.styles import Font
import re
from kinopoisk_parser import KinopoiskParser
import sys
import time
from threading import Thread  # для рисования progressbar
from db import DataBase, CONNECT_FILE, DATA_FILE


# TODO: обновить файлы requirements.txt c необходимыми библиотеками для проектов (pip freeze > requirements.txt)


# Класс для фильмов
class Film:
    """Класс занимается переименованием, копированием фильмов, записью в xlsx данных.
    """
    data_file = DATA_FILE  # файл для хранения данных о фильмах и сериалах
    # data_file = r'C:\install\Films.xlsx'  # файл для хранения данных о фильмах и сериалах

    def __init__(self, name: str, path: str, serv: str, chief: str):
        """
        :param name: полное имя файла вида жанр_год_название.ext
        :param path: путь до name
        :param serv: имя диска для записи на сервер вида "C:"
        :param chief: имя диска для просмотра вида "D:"
        """
        self.fullname = name  # жанр_год_название.ext
        self.path = path
        self.serv_disk = serv
        self.chief_disk = chief
        self.genre = None
        self.year = None
        self.id = None  # id на Кинопоиске
        self.name = None  # название.ext
        self.short_name = None  # жанр_год_название
        self.clear_name = None  # название
        self.season = None
        self.ext = None  # Отрезанное расширение файла .mkv
        if self.fullname:
            self.file_size = os.path.getsize(self.path + '\\' + self.fullname) / 1024 / 1024 / 1024
        else:
            self.file_size = None

    def split_fullname(self):
        """Разбиваем полное название файла вида жанр_год_название.формат на части"""
        self.genre, self.year, self.name = self.fullname.split('_')
        self.genre = self.genre.capitalize()  # Первый символ строки большой, остальные маленькие
        self.clear_name = os.path.splitext(self.name)[0]

    def copy_to_servdisk(self):
        """Запись на серверный диск.

        :return: True -- если копирование удалось, иначе вывод строки об ошибке копирования.
        """
        if self.file_size > 10.0:  # Размер файла больше 10 ГБ
            print(f"КОПИРОВАНИЕ НЕ УДАЛОСЬ. [Размер файла {self.fullname} больше 10Гб]")
            return
        if self.check_subtitr_film():  # Проверяем есть ли версия без субтитров
            print(f"КОПИРОВАНИЕ НЕ УДАЛОСЬ. [Имеется измененная версия {self.fullname} в {self.path}]")
            return
        self.newname_for_copy()  # Переименовываем, если заканчивается на (1)
        if self.copy_film(self.serv_disk + '/Фильмы/' + self.genre):
            with open(self.serv_disk + '/Фильмы/' + self.genre + '/' + self.genre + '.doc', 'a+') as film_genre:
                film_genre.write(self.clear_name + '\t' + self.year + '\n')  # запись в файл жанров
            self.write_films_xlsx()  # запись в Excel
            return True

    def copy_to_chiefdisk(self):
        """Запись на диск для просмотра."""
        if self.check_subtitr_film():  # Проверяем есть ли версия без субтитров
            print(f"КОПИРОВАНИЕ НЕ УДАЛОСЬ. [Имеется измененная версия {self.fullname} в {self.path}]")
            return
        self.newname_for_copy()  # Переименовываем, если заканчивается на (1)
        self.copy_film(self.chief_disk)

    def copy_film(self, destination):
        """Копирование файла.

        В случае невозможности копирования файла выводится поясняеющее сообщение.
        Для отображения progressbar при копировании используется еще один поток, сравниваюищй размеры конечного и
        исходного файлов.

        :param destination: путь копирования файла
        :return bool: маркер удачного копирования файла
        """
        disk = destination[:2]
        free = psutil.disk_usage(disk).free / (1024 * 1024 * 1024)  # свободное место на диске
        self.file_size = os.path.getsize(self.path + '\\' + self.fullname) / 1024 / 1024 / 1024  # размер файла
        if self.file_size < free:  # Если фильм помещается на диск
            if os.path.exists(destination) is False:
                os.makedirs(destination)  # создаем директорию Жанр к будущему файлу
            if os.path.exists(destination + '/' + self.name):  # Если файл уже записан, то не копируем
                print(f"КОПИРОВАНИЕ НЕ УДАЛОСЬ. [{self.fullname} уже есть в {destination}]")
                return False
            else:
                Thread(name='ProgressBar', target=self.progress_bar, args=(self.path + '\\' + self.fullname,
                                                                           destination + '\\' + self.name)).start()
                shutil.copy(self.path + '/' + self.fullname, destination + '/' + self.name)
                # Для вывода строки 100% копирования файла
                # if os.path.getsize(self.path + '/' + self.fullname) == os.path.getsize(destination + '/' + self.name):
                #     sys.stdout.write('\r')
                #     sys.stdout.write(f"{int(self.file_size*1024)} / {int(self.file_size*1024)} Mb   [ "
                #                      + "{:20s}".format('█' * 20) + " ]   100% ")
                #     sys.stdout.flush()
            print('\r' + self.path + '\\' + self.fullname + ' СКОПИРОВАНО В ' + destination + '/' + self.name)
            return True

    def check_subtitr_film(self):
        """Проверка на существование дубликата без субтитров и посторонних дорожек.

        Дубликат появляется путем обработки(конвертации, удаления дорожек, субтитров) исходника и сохранения в той же
        директории с именем "имя (1)".

        :return bool: наличие или отсутствие такого файла
        """
        self.split_fullname()
        self.short_name, self.ext = os.path.splitext(self.fullname)
        if (os.path.exists(self.path + '/' + self.short_name + '(1)' + self.ext)) or \
                (os.path.exists(self.path + '/' + self.short_name + ' (1)' + self.ext)):
            return True
        else:
            return False

    def newname_for_copy(self):
        """ Изменение имени фильмов, оканчивающихся на (1).

        Изменяется только имя для копирования, сам файл не переименовывается.
        Это обработка копии без субтитров и лишних дорожек. Именно их и записываем на диски. Оригинал не пишем.
        name  = os.path.splitext(spl[2]) == [название, расширение]
        """
        if self.clear_name.endswith('(1)'):  # Если название заканчивается на (1), то обрезаем (1)
            self.clear_name = self.clear_name[:-3].rstrip()
            self.name = self.clear_name + self.ext

    def write_films_xlsx(self):
        """Запись в файл Films.xlsx

        Данные записываются в таблицу в следующие столбцы:
        Сериал - Name | Type | Season | Kinopoisk ID
        Фильм  - Name | Type | Genre | Year

        :return True: в случае успешной записи
        """
        wb = openpyxl.load_workbook(filename=Film.data_file)
        ws = wb.active
        font = Font(name='Times New Roman',
                    size=12,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

        # ws["A" + str(ws.max_row + 1)] = self.clear_name  # Без использования self._letter_plus_row()
        cell_name = self._letter_plus_row(ws, 'Name')
        ws[cell_name] = self.clear_name
        ws[cell_name].font = font
        ws[self._letter_plus_row(ws, 'ID')] = str(ws.max_row)
        ws[self._letter_plus_row(ws, 'ID')].font = font
        if self.season:
            ws[self._letter_plus_row(ws, 'Season')] = 'Сезон ' + str(self.season)
            ws[self._letter_plus_row(ws, 'Season')].font = font
            ws[self._letter_plus_row(ws, 'Type')] = 'Сериал'
        else:
            ws[self._letter_plus_row(ws, 'Type')] = 'Фильм'
        ws[self._letter_plus_row(ws, 'Type')].font = font
        if self.genre: ws[self._letter_plus_row(ws, 'Genre')] = self.genre
        if self.year: ws[self._letter_plus_row(ws, 'Year')] = self.year
        if self.id:
            ws[self._letter_plus_row(ws, 'Kinopoisk ID')] = self.id
            ws[self._letter_plus_row(ws, 'Kinopoisk ID')].font = font
        wb.save(Film.data_file)
        wb.close()
        return True

    def _letter_plus_row(self, sheet, column):
        """Получить строку "столбец + последняя строка"

        Возвращает строку столбец плюс номер максимальной строки в таблице, но, если столбец Name,
        то возвращает номер следующей за максимальной строкой, т.к. в этом случае это новая запись.

        :param sheet: рабочий лист
        :param column: Название столбца
        :return: str ~ "D34"
        """
        if column != 'Name':
            return self._find_column_letter(sheet, column) + str(sheet.max_row)
        else:
            return self._find_column_letter(sheet, column) + str(sheet.max_row + 1)

    @staticmethod
    def _find_column_letter(sheet, column):
        """Находит букву столбца по его имени.

        Поиск производится в первой строке таблицы, т.к. именно там прописаны названия столбцов.
        Если такого столбца нет, то возвращает букву следующего за максимальным.

        :param sheet: лист Excel
        :param column: Название нужного столбца
        :return: letter: str
        """
        for cell in sheet[1]:
            if cell.value == column:
                return cell.column_letter  # column
        return sheet.cell(sheet.max_row, sheet.max_column + 1).column_letter  # column
        # letter = (cell.column_letter for cell in sheet[1] if cell.value == column)
        # if isinstance(next(letter), str):
        #     return letter + str(sheet.max_row)

    def progress_bar(self, source, dest):
        """ Отрисовывание прогресса копирования в консоли

        Никогда не отрисует 100%, т.к. цикл прервется при совпадении размеров файлов.

        :param source: Файл-источник
        :param dest: Файл назначения
        :return: sys.stdout.write()
        """
        time.sleep(0.02)
        if os.path.exists(dest):
            while os.path.getsize(source) != os.path.getsize(dest):
                sys.stdout.write('\r')
                percent = int(
                    (float(os.path.getsize(dest)) / float(os.path.getsize(source))) * 100)
                steps = int(percent / 5)
                source_size = int(os.path.getsize(dest) / 1024000)
                sys.stdout.write(f"{source_size} / {int(self.file_size * 1024)} Mb   [ "
                                 + "{:20s}".format('█' * steps) + f" ]   {percent}%  Копирование {source}")
                sys.stdout.flush()
                time.sleep(0.05)


# Класс для сериалов
class Serial(Film):
    """Класс-наследник класса Films. Занимается обработкой сериалов.

    В отличии от класса Films, ему передается не название фильма, а список серий сериала из текущей папки.
    """

    def __init__(self, files, path, serv, chief, name=None):
        """
        :param files: список серий
        :param path: путь до папки с files
        :param serv: имя диска для записи на сервер вида "C:"
        :param chief: имя диска для просомтра вида "D:"
        :param name: None
        """
        Film.__init__(self, name, path, serv, chief)
        self.season = 1
        self.files = files
        self.name = "Episode "
        self.episode_number = 1

    def name_and_season(self, disk):
        """Нахождение имени и сезона сериала по именам папок в пути.

        Проверяем различные комбинации структуры и именования папок сериала.
        Получаем имя и сезон сериала из названий папкок. Если сезон не указан ни коим образом в названиях, то
        его значение остается "1" по-умолчанию.

        #неактуально на текущий момент: Потом проверяем на наличие сезона в месте назначения и
        увеличиваем сезон, если он там есть и файлы не совпадают с исходными(закомментированный участок кода).

        :param disk: диск назначения вида "C:"
        """
        # Отрезаем полное название сериала
        if os.path.dirname(self.path) == os.getcwd():  # Если имя предыдущей директории совпадает с рабочим путем
            self.clear_name = os.path.basename(self.path)  # присваиваем имя текущего каталога
        else:  # иначе присваеваем имя предыдущего каталога (находимся в подкаталоге Номер сезона)
            self.clear_name = os.path.basename(os.path.dirname(self.path))
        # [id_serial (3 сезон)] Отрезаем id от названия
        if re.search('_', self.clear_name):
            self.id, self.clear_name = re.split('_', self.clear_name)
        if not self.clear_name.istitle():  # Проверяем начинается ли с большой буквы имя сериала
            self.clear_name = self.clear_name.capitalize()
        # Получем номер сезона, если текущая директория вида "Сезон 1"
        if re.fullmatch('[С,с]езон \d{1,2}', os.path.basename(self.path)):
            self.season = re.split(' ', os.path.basename(self.path))[1]
            if self.clear_name.endswith('сезон)'):
                self.clear_name = self.clear_name[:self.clear_name.find('(')].strip()
            return
        if self.clear_name.endswith('сезон)'):
            self.season = self.clear_name[self.clear_name.find('(') + 1:-6].strip()
            if not self.season.isdigit():
                self.season = self.season[0]
                self.season = int(self.season)
            self.clear_name = self.clear_name[:self.clear_name.find('(')].strip()

        # while os.path.exists(disk + '/' + self.clear_name + '/Сезон ' + str(self.season)):
        #     if (self.files != [] and not os.path.exists(disk + '/' + self.clear_name + '/Сезон '
        #                                 + str(self.season) + '/Episode 1' + os.path.splitext(self.files[0])[1])) or (
        #             self.files != [] and os.path.getsize(self.path + '\\' + self.files[0]) != os.path.getsize(disk
        #         + '/' + self.clear_name + '/Сезон ' + str(self.season) + '/Episode 1'
        #         + os.path.splitext(self.files[0])[1])):
        #         self.season += 1
        #     else:
        #         self.season = str(self.season)
        #         break

    def find_season(self):
        """Поиск вхождения подстроки с сезоном и эпизодом (S01E01 или s01e01) в названии файла.

        Теоретически более точный способ получить правильные сезон и номер серии, чем метод name_and_season()
        Дополнительно делает смещение нумерации серий, пропуская 13 (личная хотелка).

        :return bool: маркер получения номера сезона и эпизода из названия серии
        """
        result = re.search('[s,S][0-9]{1,3}[e,E][0-9]{1,3}', self.fullname)
        if result:  # отрезаем S от результата => делим строку по букве E и получаем сезон и эпизод
            self.season, self.episode_number = re.split('[e, E]', result.group(0)[1:])
            self.episode_number = int(self.episode_number)
            self.season = str(int(self.season))  # Первращение туда сюда убирает лишние нули в начале номера
            if self.episode_number > 12:
                self.episode_number += 1
            return True
        return False

    def copy_to_chiefdisk(self):
        """Копирование на диск для просмотра."""
        self.episode_number = 1  # без этого при вызове функции копирования на другой диск номера эпизодов продолжатся
        self.files.sort(key=len)
        self.name_and_season(self.chief_disk)
        destination = f"{self.chief_disk}/{self.clear_name}/Сезон {self.season}"
        for self.fullname in self.files:
            self.ext = os.path.splitext(self.fullname)[1]
            if self.find_season():  # Если есть подстрока с сезоном, то лучше переделать путь назначения
                destination = f"{self.chief_disk}/{self.clear_name}/Сезон {self.season}"
            self.name = "Episode " + str(self.episode_number) + self.ext
            if self.copy_film(destination):
                self.episode_number += 1
                if self.episode_number == 13:  # Второй метод увеличения номера эпизодов для случая, когда нет
                    self.episode_number += 1  # подстроки S01E01 в названии файла
            else:
                return print(f" КОПИРОВАНИЕ {self.clear_name} Cезон {self.season} в {destination} НЕ УДАЛОСЬ")

    def copy_to_servdisk(self):
        """Копирование на серверный диск.

        :return bool: маркер успешного копирования
        """
        self.episode_number = 1  # без этого при вызове функции копирования на другой диск номера эпизодов продолжатся
        self.files.sort(key=len)
        self.name_and_season(self.serv_disk + '/Сериалы')  # добавляем /Сериалы, т.к.на серв диске другой итоговый путь
        destination = f"{self.serv_disk}/Сериалы/{self.clear_name}/Сезон {self.season}"
        for self.fullname in self.files:
            self.ext = os.path.splitext(self.fullname)[1]
            if self.find_season():  # Если есть подстрока с сезоном, то лучше передалать путь назначения
                destination = f"{self.serv_disk}/Сериалы/{self.clear_name}/Сезон {self.season}"
            self.name = "Episode " + str(self.episode_number) + self.ext
            if self.copy_film(destination):
                self.episode_number += 1
                if self.episode_number == 13:
                    self.episode_number += 1
            else:
                print(f" КОПИРОВАНИЕ {self.clear_name} Cезон {self.season} в {destination} НЕ УДАЛОСЬ")
                return False
        self.write_films_xlsx()
        return True


def path_existence_check(path):
    """Проверка существования пути и смена рабочей директории.

    :param path: путь для проверки
    :return bool: маркер смены пути
    """
    if os.path.exists(path):
        os.chdir(path)
        return True
    else:
        print(f'Пути {path} не существует')
        return False


if __name__ == "__main__":
    arc_disk = input('Введите букву диска с архивом: ').upper() + ':'
    chief_disk = input('Введите букву диска для просмотра: ').upper() + ':'
    serv_disk = input('Введите букву диска для записи на видеосервер: ').upper() + ':'
    work_paths = (os.path.join(arc_disk, r'\Convert'), os.path.join(arc_disk, r'\New'))
    with KinopoiskParser() as browser:
        browser.open_selenium()
        for work_path in work_paths:  # проверяем существуют ли указанные рабочие пути
            print("=" * 150)
            if path_existence_check(work_path):
                for adress, dirs, files in os.walk(work_path):
                    if adress == work_path:  # если находимся в корне рабочего пути, то обрабатываем все фильмы
                        for file in files:
                            film = Film(file, adress, serv_disk, chief_disk)
                            if film.copy_to_servdisk():
                                browser.write_data()  # находим и дописываем данные с Кинопоиска
                                with DataBase(CONNECT_FILE) as base:  # запись в Postgresql
                                    base.query()
                            film.copy_to_chiefdisk()
                            print("=" * 150)
                    elif not dirs:  # если не в корне - значит это сериал. Запись происходит когда уже в папке с сериями
                        serial = Serial(files, adress, serv_disk, chief_disk)
                        if serial.copy_to_servdisk():
                            browser.write_data()  # находим и дописываем данные с Кинопоиска
                            with DataBase(CONNECT_FILE) as base:  # запись в Postgresql
                                base.query()
                        print("-" * 150)
                        serial.copy_to_chiefdisk()
                        print("=" * 150)
            else:
                continue
