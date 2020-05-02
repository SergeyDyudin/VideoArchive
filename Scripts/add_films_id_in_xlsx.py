import selenium
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
from kinopoisk_parser import KinopoiskParser
import time
import os

"""Этот вариант скрипта не банит Кинопоиск при массовом заполнении, в отличии от варианта add_data_in_xlsx.py 
    с kinopoisk_parser.py.
    Запросы через requests или отдельный запуск selenium для каждого фильма банятся кинопоиском.
    В данном варианте с одной сессии Selenium заполняется вся таблица и бана не происходит.
     ____________________________________________________________________________________________________________
    | Name | Type | Season | genre | Year | Kinopoisk ID | IMDB | Kinopoisk | Country | Time | Director | Actors | 
     ------+------+--------+-------+------+--------------+------+-----------+---------+------+----------+--------
    |      |      |        |       |      |              |      |           |         |      |          |        |
"""


url = 'https://www.kinopoisk.ru/'
browser = KinopoiskParser()
browser.open_selenium()
wb = openpyxl.load_workbook(filename='C:/install/Films.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[1].value == 'Сериал': continue  # пока не трогаем сериалы
    if row[5].value: continue
    name_film = str(row[0].value).lower()
    year_film = str(row[4].value)
    search_form = browser.browser.find_element_by_name('kp_query')
    search_form.clear()
    search_form.send_keys(name_film)
    search_form.submit()
    time.sleep(6)
    WebDriverWait(browser.browser, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'name')))
    results = browser.browser.find_elements_by_class_name('name')
    for result in results:
        try:
            year = ''
            year = result.find_element_by_class_name('year').text
        except:
            continue  # break
        if '-' in year:
            continue
        """ else:
            year = int(year)"""
        name = result.find_element_by_tag_name('a').text
        # В таблице во всех названиях ':' заменено на '.', т.к. названия файлов не позволяют ставить ':'
        name = name.replace(':', '.')
        # в названиях на Кинопоиске иногда попадаются специальные символы многоточия, которые дают False в сравнении имен
        name = name.replace(chr(8230), '...')
        name_film = name_film.replace(chr(8230), '...')
        # в названиях на Кинопоиске иногда попадаются символ неразрывного пробела, который дает False в сравнении имен
        name = name.replace(chr(160), chr(32))
        if (name_film == name.lower()) and (year_film == year):
            result.find_element_by_tag_name('a').click()
            time.sleep(10)
            data = {}
            data['id_kinopoisk'] = os.path.basename(os.path.dirname(browser.browser.current_url))
            try:
                row[5].value = data["id_kinopoisk"]
            except KeyError:
                print('Нет значения id_kinopoisk')
            wb.save(r'C:\install\Films.xlsx')
            print(f'Получены и записаны данные для {name_film}')
            break
wb.close()
browser.close_selenium()