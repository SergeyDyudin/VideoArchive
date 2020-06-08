import openpyxl
from openpyxl.styles import Font
import time
import re
from kinopoisk_parser import KinopoiskParser
from bs4 import BeautifulSoup

browser = KinopoiskParser()
browser.open_selenium()
wb = openpyxl.load_workbook(filename='C:/install/Films.xlsx')
ws = wb.active
font = Font(name='Times New Roman',
                    size=12,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
for row in ws.iter_rows(min_row=2235, max_row=ws.max_row):
    result = {}
    if row[1].value == 'Фильм': continue
    id_film = str(row[5].value)
    season = int(re.split(' ', row[2].value)[1])
    url = f'https://www.kinopoisk.ru/film/{id_film}'
    browser.browser.get(url)
    time.sleep(5)
    try:
        element = browser.browser.find_element_by_class_name('table-col-years__seasons')
        browser.browser.execute_script("arguments[0].click();", element)
        time.sleep(5)
        get = browser.browser.page_source
        soup = BeautifulSoup(get, features="html.parser")
        years = []
        items = soup.find('div', {'class', 'shadow'}).find_all('a', {'class', 'all'})
        for i in items:
            if '#y' in i['href']:
                years += i.contents
        result["year"] = years
        result["year"] = result["year"][season - 1]
        row[4].value = result["year"]
        row[4].font = font
        wb.save('C:/install/Films.xlsx')
        print(f'Даты для {row[0].value} записаны')
    except Exception as exc:
        print(exc)
        print(f'Даты сезонов не получилось достать. {row[0].value}')
wb.close()
browser.close_selenium()