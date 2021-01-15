import requests
import openpyxl
from pathlib import Path
import os


if os.path.exists('/home/zs-content-02-usr/projects/django/films_site/main/static/main/img/icons'):  # Ubuntu-serv
    PATH_TO_ICONS = '/home/zs-content-02-usr/projects/django/films_site/main/static/main/img/icons'
elif os.path.exists('/Users/sergeydyudin/Documents/PycharmProjects/films_site/main/static/main/img/icons'):
    PATH_TO_ICONS = '/Users/sergeydyudin/Documents/PycharmProjects/films_site/main/static/main/img/icons'
else:
    raise 'Добавьте путь к папке иконками на этом PC'
def get_film_icon():
    file_xlsx = Path(__file__).parent.parent.joinpath('database').joinpath('Films.xlsx')

    wb = openpyxl.load_workbook(filename=file_xlsx)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        id_xlsx = str(row[0].value)
        id_film = str(row[5].value)
        url = f'https://www.kinopoisk.ru/images/sm_film/{id_film}.jpg'
        req = requests.get(url)
        with open(PATH_TO_ICONS + f'/{id_film}_icon.jpg', 'wb') as icon_file:
            icon_file.write(req.content)
            print(f'Загружена иконка для фильма {id_film} {id_xlsx}')
    wb.close()

if __name__ == "__main__":
    get_film_icon()