from db import DataBase, get_xlsx_path, get_dbauth_file_path
import openpyxl


file_xlsx = get_xlsx_path()
connect_file = get_dbauth_file_path()

wb = openpyxl.load_workbook(filename=file_xlsx)
ws = wb.active

db = DataBase(connect_file)
for row in range(2, ws.max_row+1):
    description = ws.cell(row=row, column=db._find_column(ws, 'Description')).value
    films_id = ws.cell(row=row, column=db._find_column(ws, 'ID')).value
    sql_request = f"UPDATE films SET description=%s WHERE id=%s;"
    param = (description, films_id)
    db.request(sql_request, param)
wb.close()
db.close()
