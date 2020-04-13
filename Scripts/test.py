import selenium
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
from Scripts.kinopoisk_parser import KinopoiskParser
import time

url = 'https://www.kinopoisk.ru/'
browser = webdriver.Firefox()  # firefox_profile=r'C:\Users\video\AppData\Roaming\Mozilla\Firefox\Profiles\h3qugs8n.Kinopisk')
browser.get(url)
# search_form = browser.find_element_by_name('kp_query')

wb = openpyxl.load_workbook(filename='C:/install/Films.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=1976, max_row=2488):# max_row=ws.max_row):
    if row[1].value == 'Сериал': continue  # пока не трогаем сериалы
    if not row[4].value: continue  # пока не трогаем первую плохо заполеннную часть таблицы, где нет Year
    if row[11].value: continue  # пропускаем уже заполеннные строки
    name_film = row[0].value.lower()
    year_film = str(row[4].value)
    search_form = browser.find_element_by_name('kp_query')
    search_form.clear()
    search_form.send_keys(name_film)
    search_form.submit()
    time.sleep(6)
    WebDriverWait(browser,10).until(EC.presence_of_all_elements_located((By.CLASS_NAME,'name')))
    # WebDriverWait(browser, 10).until(EC.presence_of_all_elements_located((By.PARTIAL_LINK_TEXT, name_film.capitalize())))
    # get = browser.page_source
    # soup = BeautifulSoup(get, features="html.parser")
    # results = soup.find_all('p', {'class': 'name'})
    results = browser.find_elements_by_class_name('name')
    for result in results:
        try:
            year = result.find_element_by_class_name('year').text
        except:
            break
        if '-' in year:
            continue
        """ else:
            year = int(year)"""
        name = result.find_element_by_tag_name('a').text
        name = name.replace(':', '.')
        if (name_film == name.lower()) and (year_film == year):
            result.find_element_by_tag_name('a').click()
            break
    try:
        WebDriverWait(browser,10).until(EC.presence_of_element_located((By.CLASS_NAME,'info')))
        WebDriverWait(browser,10).until(EC.presence_of_element_located((By.ID,'actorList')))
        get = browser.page_source
        data = KinopoiskParser.get_info(get)
        if (row[0].value == data['film_name'].replace(':', '.')) and (row[4].value == int(data['year'])):
            try:
                row[3].value = data["jenre"]
            except KeyError:
                print('Нет значения жанр')
            try:
                row[5].value = data["id_kinopoisk"]
            except KeyError:
                print('Нет значения id_kinopoisk')
            try:
                row[6].value = data["imdb"]
            except KeyError:
                print('Нет значения IMDB')
            try:
                row[7].value = data["kinopoisk"]
            except KeyError:
                print('Нет значения Kinopoisk')
            try:
                row[8].value = data["country"]
            except KeyError:
                print('Нет значения country')
            try:
                row[9].value = data["time"]
            except KeyError:
                print('Нет значения Time')
            try:
                row[10].value = data["director"]
            except KeyError:
                print('Нет значения Director')
            try:
                actors = ''
                for key, value in data['actors'].items():
                    actors += value + ", "
                row[11].value = actors[:-2]
            except KeyError:
                print('Нет значения Actors')
            wb.save(r'C:\install\Films.xlsx')
            print(f'Получены и записаны данные для {name_film}')
    except selenium.common.exceptions.TimeoutException:
        print('Что-то пошло не так с этим фильмом')
browser.close()