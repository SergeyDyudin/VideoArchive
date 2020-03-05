import openpyxl
import re

wb = openpyxl.load_workbook(filename='C:/install/Films.xlsx')
ws = wb.active
for i in range(1, ws.max_row):
    string = ws["B" + str(i)].value
    #if not string:
    #    string = "None"
    if ws["B" + str(i)].value and re.search('[С, с]езон[ы]? \d{1,2}-\d{1,2}', ws["B" + str(i)].value):

        print(f'{i} => {ws["B" + str(i)].value}')
        #print(re.split(" ", ws["B" + str(i)].value)[1])
        result = re.split("-", re.split(" ", ws["B" + str(i)].value)[1])
        for s in range(int(result[0]), int(result[1]) + 1):
            print(s)
            cell = str(ws.max_row + 1)
            if s == int(result[0]):
                ws["B" + str(i)] = f"Сезон {s}"
                print(f'{str(i)}: {ws["A" + str(i)].value} => {ws["B" + str(i)].value}')
            else:
                ws["A" + cell] =  ws["A" + str(i)].value
                ws["B" + cell] = f"Сезон {s}"
                print(f'{cell}: {ws["A" + cell].value} => {ws["B" + cell].value}')

wb.save(r'C:\install\Films.xlsx')
