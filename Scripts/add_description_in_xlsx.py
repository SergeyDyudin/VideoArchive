import openpyxl
from openpyxl.styles import Font
import time
from kinopoisk_parser import KinopoiskParser
from bs4 import BeautifulSoup
from selenium import webdriver
from pathlib import Path

browser = KinopoiskParser()
driver_path = str(Path(__file__).parent.parent.joinpath('web-drivers').joinpath('chromedriver'))
file_xlsx = Path(__file__).parent.parent.joinpath('database').joinpath('Films.xlsx')

browser.browser = webdriver.Chrome(driver_path)
wb = openpyxl.load_workbook(filename=file_xlsx)
ws = wb.active
font = Font(name='Times New Roman',
                    size=12,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
for row in ws.iter_rows(min_row=2, max_row=4):
# for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    result = {}
    id_film = str(row[5].value)
    url = f'https://www.kinopoisk.ru/film/{id_film}'
    browser.browser.get(url)
    time.sleep(5)
    try:
        get = browser.browser.page_source
        soup = BeautifulSoup(get, features="html.parser")
        description = ''
        items = soup.find('div', {'class', 'styles_filmSynopsis__zLClu'}).find_all('p')
        for i in items:
            description += i.contents[0]
        row[13].value = description
        row[13].font = font
        wb.save(file_xlsx)
        print(f'Описание для {row[0].value} записаны')
    except Exception as exc:
        print(exc)
        print(f'Описание не получилось достать. {row[0].value}')
wb.close()
browser.close_selenium()