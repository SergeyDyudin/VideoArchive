from Scripts.kinopoisk_parser import KinopoiskParser
import openpyxl
import time

""" Данный скрипт заполняет таблицу в Films.xlsx данными из Кинопоиска
    В таблице заполнены только поля Name, Year и Type(Фильм/Сериал). По этим полям находятся данные на Кинопоиске. 
"""

"""UPD: Данный скрипт отрабатывает, если количество запросов 3-4, иначе Кинопоиск увеличивает время ответа на запросы
"""

"""name_film = input('Введите имя фильма: ')
name_film = name_film.lower()
year_film = input('Введите год: ')
film = KinopoiskParser(name=name_film, year=year_film)
film.find_film_id()
data = film.get_from_kinopoisk_with_id()"""
# data = KinopoiskParser().get_from_file('C:/Users/Kenobi/Desktop/Нокдаун.html')
# print(data)

wb = openpyxl.load_workbook(filename='C:/install/Films.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=409, max_row=412):  # max_row=ws.max_row):
    if row[1].value == 'Сериал': continue  # пока не трогаем сериалы
    if not row[4].value: continue  # пока не трогаем первую плохо заполеннную часть таблицы
    name_film = row[0].value.lower()
    year_film = str(row[4].value)
    film = KinopoiskParser(name=name_film, year=year_film)
    id = film.find_film_id()
    if not id:
        time.sleep(10)
        continue
    data = film.get_from_kinopoisk_with_id()
    if (row[0].value == data['film_name'].replace(':','.')) and (row[4].value == int(data['year'])):
        try:
            row[3].value = data["jenre"]
        except KeyError:
            print('Нет значения жанр')
        try:
            row[5].value = data["id_kinopoisk"]
        except KeyError:
            print('Нет значения id_kinopoisk')
        try:
            row[6].value = data["imdb"]
        except KeyError:
            print('Нет значения IMDB')
        try:
            row[7].value = data["kinopoisk"]
        except KeyError:
            print('Нет значения Kinopoisk')
        try:
            row[8].value = data["country"]
        except KeyError:
            print('Нет значения country')
        try:
            row[9].value = data["time"]
        except KeyError:
            print('Нет значения Time')
        try:
            row[10].value = data["director"]
        except KeyError:
            print('Нет значения Director')
        try:
            actors = ''
            for key, value in data['actors'].items():
                actors += value + ", "
            row[11].value = actors[:-2]
        except KeyError:
            print('Нет значения Actors')
    print(f'Получены и записаны данные для {name_film}')
    wb.save(r'C:\install\Films.xlsx')
    time.sleep(10)  # задержка между запросами, чтобы не банил Кинопоиск
wb.save(r'C:\install\Films.xlsx')
wb.close()
