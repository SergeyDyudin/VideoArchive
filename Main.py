import os
import shutil
import psutil
import openpyxl
import xlrd

arc_disk = input('Введите букву диска с архивом: ').upper() + ':'
chief_disk = input('Введите букву диска для записи шефу: ').upper() + ':'
serv_disk = input('Введите букву диска для записи на видеосервер: ').upper() + ':'


"""Запись на диск для видеосервера
"""

work_path = (arc_disk + '/Convert', arc_disk + '/New')
for work_path_i in work_path:
   if os.path.exists(work_path_i): os.chdir(work_path_i)
   else:
       print(work_path_i + ' не существует')
       continue
   s = 1 #номер сезона
   #DISK = "C:"
   #all_size = psutil.disk_usage(DISK).total/(1024*1024*1024)#размер диска
   for adress, dirs, files in os.walk(os.getcwd()):#получаем (текущая директория, подкаталоги, файлы)
      #if (len(dirs) != []) and (files == []): s = 1
      s = 1 #номер сезона
      if adress == os.getcwd():#проверка фильм или сериал(фильмы лежат в корне, сериалы в поддиректориях)
          #if os.path.exists(serv_disk + '/Фильмы/') == False: os.mkdir(serv_disk + '/Фильмы/')
          for f in files:
              new_name = '' 
              file_size = os.path.getsize(adress + '\\' + f)/1024/1024/1024#размер файла
              free = psutil.disk_usage(serv_disk).free/(1024*1024*1024)#свободное место на диске
              if (file_size < free) and (file_size < 9.0):#если фильм помещается на диск и его размер меньше 9 ГБ
                  spl = f.split('_')#Разбиваем длинное название на [жанр, год, имя]
                  if (os.path.exists(arc_disk + '/New/' + os.path.splitext(spl[2])[0] + '(1)' + os.path.splitext(spl[2])[1])) or (os.path.exists(arc_disk + '/New/' + os.path.splitext(spl[2])[0] + ' (1)' + os.path.splitext(spl[2])[1])):
                     continue # проверка на существование дубликата без субтитров и посторонних дорожек на архивном диске
                  spl[0] = spl[0].capitalize()#Первый символ строки большой, остальные маленькие
                  if os.path.exists(serv_disk + '/Фильмы/'+spl[0]+'/'+spl[2]): #если файл уже есть на диске, то пропускаем его
                     continue
                  if os.path.exists(serv_disk + '/Фильмы/'+spl[0]) == False: os.makedirs(serv_disk + '/Фильмы/'+spl[0])#создаем директории к будущему файлу
                  if (os.path.splitext(spl[2])[0][-1] == ')') and (os.path.splitext(spl[2])[0][-2] == '1') and (os.path.splitext(spl[2])[0][-3] == '('):#Если это копия (1), то копируем с обычным именем
                      for i in range(0, len(os.path.splitext(spl[2])[0])-4):
                          new_name = new_name + os.path.splitext(spl[2])[0][i]
                      if os.path.exists(serv_disk + '/' + new_name + os.path.splitext(spl[2])[1]):
                         continue
                      z = shutil.copy(f, serv_disk +  '/Фильмы/'+spl[0] + '/' + new_name + os.path.splitext(spl[2])[1])           
                  else: z = shutil.copy(f, serv_disk +  '/Фильмы/' +spl[0] + '/'+spl[2])
                  print(adress + '\\' + f + ' СКОПИРОВАНО В ' + z)
                  film_jenre = open(serv_disk + '/Фильмы/'+spl[0]+'/' + spl[0] + '.doc', 'a+')#запись в файл жанров
                  if new_name != '': film_jenre.write(new_name +'\t' + spl[1] + '\n')
                  else: film_jenre.write(os.path.splitext(spl[2])[0] +'\t' + spl[1] + '\n')
                  film_jenre.close()
                  wb = openpyxl.load_workbook(filename='C:/install/Films.xlsx')#запись в общую базу фильмов
                  ws = wb.active
                  rb = xlrd.open_workbook('C:\install\Films.xlsx')
                  sheet = rb.sheet_by_index(0)
                  if new_name != '': ws["A" + str(sheet.nrows + 1)] = new_name
                  else: ws["A" + str(sheet.nrows +1)] = os.path.splitext(spl[2])[0]
                  ws["B" + str(sheet.nrows +1)] = spl[0]
                  ws["C" + str(sheet.nrows +1)] = spl[1]
                  wb.save('C:\install\Films.xlsx')
        
                  # if work_path_i == arc_disk + '/Convert': os.remove(f)
      else:
         if os.path.split(adress)[0] == os.getcwd():
            name_serial = adress.split('\\')[-1].capitalize()#отрезаем название сериала из пути
            wb = openpyxl.load_workbook(filename='C:/install/Films.xlsx')#запись в общую базу фильмов
            ws = wb.active
            rb = xlrd.open_workbook('C:\install\Films.xlsx')
            sheet = rb.sheet_by_index(0)
            ws["A" + str(sheet.nrows +1)] = name_serial
            ws["B" + str(sheet.nrows +1)] = 'Сезон ' + str(s) 
            wb.save('C:\install\Films.xlsx')
         #if (os.path.exists(serv_disk + '/Сериалы/' + '/' + name_serial) == False) and (dirs != 0): os.mkdir(serv_disk + '/Сериалы/'+ '/' + name_serial)
         while os.path.exists(serv_disk + '/Сериалы/' + name_serial + '/Season ' + str(s)):
            if files != [] and os.path.getsize(adress + '\\' + sorted(files)[0]) != os.path.getsize(serv_disk + '/Сериалы/' + name_serial + '/Season ' + str(s) + '/Episode 1' + os.path.splitext(sorted(files)[0])[1]):
               s += 1
            else: break
         e = 1 #номер эпизода
         for f in sorted(files):
               if os.path.exists(serv_disk + '/Сериалы/' + '/' + name_serial + '/Season ' + str(s)) == False: os.makedirs(serv_disk + '/Сериалы/'+ '/' + name_serial + '/Season ' + str(s))#создаем директории к будущему файлу
               file_size = os.path.getsize(adress)/1024/1024/1024#размер файла
               free = psutil.disk_usage(serv_disk).free/(1024*1024*1024)
               if (file_size < free) and (file_size < 9.0):
                   #g = os.path.exists(serv_disk + '/Сериалы/'+ name_serial + '/Season ' + str(s)+ '/' + 'Episode ' + str(e))
                   if os.path.exists(serv_disk + '/Сериалы/'+ name_serial + '/Season ' + str(s)+ '/' + 'Episode ' + str(e) + os.path.splitext(f)[1]):
                      e += 1
                      if e == 13:
                         e += 1
                      continue
                   z = shutil.copy(adress + '/' + f, serv_disk + '/Сериалы/'+ name_serial + '/Season ' + str(s)+'/'+ 'Episode ' + str(e) + os.path.splitext(f)[1])
                   e += 1
                   if e == 13:
                       e += 1
                   print(adress + '\\' + f + ' СКОПИРОВАНО В ' + z)
                   # if work_path_i == arc_disk + '/Convert': os.remove(f)
         #if dirs == []: s += 1


""" Запись на диск шефу
"""

os.chdir(arc_disk + '/New')

for adress, dirs, files in os.walk(os.getcwd()):#получаем (текущая директория, подкаталоги, файлы)
   s = 1 #номер сезона
   #if (len(dirs) != []) and (files == []): s = 1
   if adress == os.getcwd():#проверка фильм или сериал(фильмы лежат в корне, сериалы в поддиректориях)
      for f in files:
         new_name= ''                                    
         file_size = os.path.getsize(adress + '\\' + f)/1024/1024/1024#размер файла
         free = psutil.disk_usage(chief_disk).free/(1024*1024*1024)#свободное место на диске        
         if file_size < free:#если фильм помещается на диск 
            spl = f.split('_')#Разбиваем длинное название на [жанр, год, имя]
            if (os.path.exists(arc_disk + '/New/' + os.path.splitext(spl[2])[0] + '(1)' + os.path.splitext(spl[2])[1])) or (os.path.exists(arc_disk + '/New/' + os.path.splitext(spl[2])[0] + ' (1)' + os.path.splitext(spl[2])[1])):
                continue # проверка на существование дубликата без субтитров и посторонних дорожек
            spl[0] = spl[0].capitalize()#Первый символ строки большой, остальные маленькие
            if os.path.exists(chief_disk + '/' + spl[2]): #если файл уже есть на диске, то пропускаем его
               continue
            if (os.path.splitext(spl[2])[0][-1] == ')') and (os.path.splitext(spl[2])[0][-2] == '1') and (os.path.splitext(spl[2])[0][-3] == '('):
                for i in range(0, len(os.path.splitext(spl[2])[0])-4):
                    new_name = new_name + os.path.splitext(spl[2])[0][i]
                if os.path.exists(chief_disk + '/' + new_name + os.path.splitext(spl[2])[1]):
                   continue
                z = shutil.copy(f, chief_disk + '/' + new_name + os.path.splitext(spl[2])[1])           
            else: z = shutil.copy(f, chief_disk + '/'+spl[2])
            print(adress + '\\' + f + ' СКОПИРОВАНО В ' + z)
                  
   else:
      if os.path.split(adress)[0] == os.getcwd():
          name_serial = adress.split('\\')[-1].capitalize()#отрезаем название сериала из пути    
      if (os.path.exists(chief_disk + '/' + name_serial) == False) and (dirs != 0): os.mkdir(chief_disk + '/' + name_serial)
      while os.path.exists(chief_disk + '/' + name_serial + '/Season ' + str(s)):
            if files != [] and os.path.getsize(adress + '\\' + sorted(files)[0]) != os.path.getsize(chief_disk + '/' + name_serial + '/Season ' + str(s) + '/Episode 1' + os.path.splitext(sorted(files)[0])[1]):
               s += 1
            else: break
      e = 1 #номер эпизода
      for f in sorted(files):
         if os.path.exists(chief_disk + '/' + name_serial + '/Season ' + str(s)) == False: os.makedirs(chief_disk + '/' + name_serial + '/Season ' + str(s))#создаем директории к будущему файлу
         file_size = os.path.getsize(adress)/1024/1024/1024#размер файла
         free = psutil.disk_usage(chief_disk).free/(1024*1024*1024)
         if file_size < free:
               g = os.path.exists(chief_disk + '/' + name_serial + '/Season ' + str(s)+ '/' + 'Episode ' + str(e))
               if os.path.exists(chief_disk + '/' + name_serial + '/Season ' + str(s)+ '/' + 'Episode ' + str(e) + os.path.splitext(f)[1]):
                  e += 1
                  if e == 13:
                     e += 1
                  continue
               z = shutil.copy(adress + '/' + f, chief_disk + '/' + name_serial + '/Season ' + str(s)+'/'+ 'Episode ' + str(e) + os.path.splitext(f)[1])
               e += 1
               if e == 13:
                  e += 1
               print(adress + '\\' + f + ' СКОПИРОВАНО В ' + z)
   #if dirs == []: s += 1
    
