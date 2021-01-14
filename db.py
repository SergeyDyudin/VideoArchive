import psycopg2
from psycopg2 import sql
import openpyxl

DATA_FILE = r'C:\install\Films.xlsx'
CONNECT_FILE = 'dbauth.txt'


class DataBase:
    """Класс для работы с базой данных
    """

    def __init__(self, conn_file=None):
        if not conn_file:
            conn_file = CONNECT_FILE
        with open(conn_file) as file:
            db_name = file.readline().rstrip()
            db_user = file.readline().rstrip()
            db_password = file.readline().rstrip()
            db_host = file.readline().rstrip()
        # Вывод данных о базе для тестирования - потом убрать
        print(f'db = {db_name}', f'user = {db_user}', f'pass = {db_password}', f'host = {db_host}', sep='\n')
        try:
            self.conn = psycopg2.connect(database=db_name, user=db_user, password=db_password, host=db_host)
            self.cur = self.conn.cursor()
            print('***Соединение с базой установлено.***')
        except Exception as err:
            print(err)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is not None:
            print('Error!', exc_type, exc_val)
        self.cur.close()
        self.conn.close()
        print('***Соединение с базой закрыто.***')

    @staticmethod
    def _find_column(sheet, name):
        """
        Поиск номера столбца по имени.

        :param sheet: рабочий лист в xlsx
        :param name: имя столбца для поиска его номера
        :return: номер столбца
        """
        for cell in sheet[1]:
            if cell.value == name:
                return cell.col_idx

    def get_data(self, num_str=None):
        """Получение словаря с данными для заполнения базы данных из файла Films.xlsx.

        :param num_str: номер строки в Films.xlsx
        :return: result
        """
        result = {}
        try:
            wb = openpyxl.load_workbook(filename=DATA_FILE)
            ws = wb.active
            if not num_str:
                num_str = ws.max_row
            result['name'] = ws.cell(row=num_str, column=self._find_column(ws, 'Name')).value
            result['type'] = ws.cell(row=num_str, column=self._find_column(ws, 'Type')).value.lower()
            result['season'] = str(ws.cell(row=num_str, column=self._find_column(ws, 'Season')).value).lower()
            result['year'] = str(ws.cell(row=num_str, column=self._find_column(ws, 'Year')).value)
            result['genre'] = ws.cell(row=num_str, column=self._find_column(ws, 'Genre')).value
            result['id_kinopoisk'] = str(ws.cell(row=num_str, column=self._find_column(ws, 'Kinopoisk ID')).value)
            result['imdb'] = ws.cell(row=num_str, column=self._find_column(ws, 'IMDB')).value
            result['kinopoisk'] = ws.cell(row=num_str, column=self._find_column(ws, 'Kinopoisk')).value
            result['country'] = ws.cell(row=num_str, column=self._find_column(ws, 'Country')).value
            result['time'] = ws.cell(row=num_str, column=self._find_column(ws, 'Time')).value
            result['director'] = ws.cell(row=num_str, column=self._find_column(ws, 'Director')).value
            result['actors'] = ws.cell(row=num_str, column=self._find_column(ws, 'Actors')).value
            if result['season']: result['season'] = result['season'].lower()
            if result['country']:
                result['country'] = result['country'].split(', ')
                for i in range(len(result['country'])):
                    result['country'][i] = result['country'][i].strip()
                # result['country'] = tuple([result['country'], ''])
            if result['director']:
                result['director'] = result['director'].split(', ')
                # result['director'] = tuple([result['director'], ''])
            if result['actors']:
                result['actors'] = result['actors'].split(', ')
                # result['actors'] = tuple([result['actors'], ''])
            result['description'] = ws.cell(row=num_str, column=self._find_column(ws, 'Description')).value
        finally:
            wb.close()
        return result

    def query(self, id_str=None):
        """Выполнение и коммит запроса для записи нового фильма в базу.

        :param id_str:int номер строки для запроса из xlsx. Если не указан, то записывается последняя строка из xlsx.
        :return:
        """
        data = self.get_data(id_str)
        # request = """
        #             INSERT INTO {table} ({field})
        #             VALUES (LOWER(%s))
        #             ON CONFLICT ({field}) DO UPDATE SET {field}=LOWER(Excluded.{field})
        #             Returning id;
        #             """
        try:
            # Заполнение всех таблиц и получение id_films
            self.cur.execute("""
                                SELECT ins_film(%(name)s, %(genre)s, %(year)s, %(type)s, %(season)s,  %(kinopoisk)s, 
                                %(imdb)s, %(id_kinopoisk)s, %(time)s, %(actors)s, %(director)s, %(country)s, 
                                %(description)s); 
                                """, data)
            id_films = self.cur.fetchone()[0]
        except psycopg2.errors.UniqueViolation:
            print(f'ERROR! Попытка повторной записи фильма в базу. Такой фильм уже есть в базе.{data["name"]}')
        except Exception as err:
            print(data["name"])
            print('Error! ', err)
            self.conn.rollback()
        else:
            print('Данные успешно добавлены в базу. Делается коммит.')
            print(f'{data["name"]} ID film = {id_films}')
            self.conn.commit()
        # print(self.cur.fetchall())

    def request(self, text):
        """
        Выполнение запроса text в базе.

        :param text: str Текст запроса в базу.
        :return: res: результат запроса
        """
        try:
            self.cur.execute(text)
            res = self.cur.fetchall()
        except Exception as e:
            print('ERROR! ', e)
        return res

    def close(self):
        self.cur.close()
        self.conn.close()


if __name__ == '__main__':
    # CONNECT_FILE = r'C:\install\dbauth.txt'
    with DataBase(CONNECT_FILE) as base:
        # print(base.get_data(12))
        base.query(17)
        # for (i, v) in base.get_data(12).items():
        #     print(i, '===', v)
