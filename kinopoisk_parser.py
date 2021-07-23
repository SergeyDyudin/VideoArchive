"""
ProxyManager: класс для получения прокси-сервера
KinopoiskParser: класс для работы с Кинопоиском
"""
from bs4 import BeautifulSoup
import requests
import selenium
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from pathlib import Path
import time
import openpyxl
from openpyxl.styles import Font
import re
import os
import sys
from db import get_dbauth_file_path


# data_file = r'C:\install\Films.xlsx'  # xlsx-файл для хранения базы данных
data_file = str(Path(__file__).parent.joinpath('database').joinpath('Films.xlsx'))
ICON_PATHS = (
    '/Users/sergeydyudin/Documents/PycharmProjects/films_site/media/main/icons/',
    'C:\\Users\\video\\Documents\\Projects\\Films_site\\media\\main\\icons\\',
    '/var/www/films_site/media/main/icons/',
    '/home/zs-content-02-usr/projects/django/films_site/media/main/icons/'
)
system = sys.platform


class ProxyManager:
    """
    Класс для получения списка доступных прокси с портами и его обновления
    """

    def __init__(self):
        self._current_proxy_index = 1
        self._proxy_list = []
        self._get_proxy_list()
        self._current_proxy = f'http://{self._proxy_list[0]}'
        self._current_proxy_s = f'https://{self._proxy_list[0]}'

    def get_proxies(self):
        """
        Получить прокси.

        :return: proxies -- словарь из http и https адресов с портом для одного прокси
        """

        """proxy_ip_with_port = self._proxy_list[self._current_proxy_index]
        self._current_proxy = f'http://{proxy_ip_with_port}'
        self._current_proxy_s = f'https://{proxy_ip_with_port}'"""
        proxies = {
            "http": self._current_proxy,
            "https": self._current_proxy_s
        }

        return proxies

    def update_proxy(self):
        """
        Взятие из списка прокси-серверов следующего значения.

        :return: self._current_proxy -- следующий http-прокси из списка
        """
        self._current_proxy_index += 1
        if self._current_proxy_index == len(self._proxy_list):
            print("Proxies are ended")
            print("Try get alternative proxy")
            proxy_ip_with_port = self._get_another_proxy()
            print("Proxy updated to " + proxy_ip_with_port)

            self._current_proxy = f'http://{proxy_ip_with_port}'
            self._current_proxy_s = f'https://{proxy_ip_with_port}'
            self._proxy_list = []
            self._proxy_list.insert(0, proxy_ip_with_port)
            self._current_proxy_index = 0
            return self._current_proxy

        proxy_ip_with_port = self._proxy_list[self._current_proxy_index]

        print("Proxy updated to " + proxy_ip_with_port)

        self._current_proxy = f'http://{proxy_ip_with_port}'
        self._current_proxy_s = f'https://{proxy_ip_with_port}'
        return self._current_proxy

    @staticmethod
    def _get_another_proxy():
        """
        Получить новый прокси-сервер из запроса к API сайта.

        https://api.getproxylist.com/proxy?protocol[]=http каждый раз при запросе выдает новый прокси

        :return: proxy -- строка вида ip:port
        """
        proxy_response = requests.get("https://api.getproxylist.com/proxy?protocol[]=http", headers={
            'Content-Type': 'application/json'
        }).json()

        ip = proxy_response['ip']
        port = proxy_response['port']
        proxy = f'{ip}:{port}'

        return proxy

    def _get_proxy_list(self):
        """
        Получить новый список прокси-серверов.

        :return: self._proxy_list - строки серверов с портами
        """
        proxy_response = requests.get("http://www.freeproxy-list.ru/api/proxy?anonymity=false&token=demo")
        # proxy_response = requests.get("http://api.foxtools.ru/v2/Proxy.txt?cp=UTF-8&lang=Auto&type=\
        # HTTPS&anonymity=None&available=Yes&free=Yes&page=1")

        self._proxy_list = proxy_response.text.split("\n")


class KinopoiskParser:
    """Получение данных о фильме или сериале.

    Класс ищет данные фильма по названию и году либо на сайте Кинопоиска, либо на сохраненной локально странице.
    Класс ищет данные сериала по ID сериала, указанному в таблице xlsx. ID получать из URL на странице сериала,
    либо параметра в коде страницы.
    На данный момент поиск актуален через использование Selenium-браузера, т.к. остальные варианты получают бан, капчу
    или увеличивающиеся задержки в выдаче ответа.
    """

    def __init__(self, id=None, name=None, year=None):
        self.id_film = id
        self.name_film = name
        self.year = year
        self.result = {}
        self.browser = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.browser.close()

    @staticmethod
    def get_info(content):
        """Получение данных фильма с помощью BeautifulSoup.

        :param content: объект, который передается в BeautifulSoup
        :return: results - словарь с данными
        """
        # text = self.read_from_file()  # Получаем данные из файла
        soup = BeautifulSoup(content, features="html.parser")
        results = {}
        actors = {}
        results['film_name'] = soup.find('span', {'class': 'styles_title__2l0HH'}).text
        # results['film_name'] = soup.find('span', {'class': 'moviename-title-wrapper'}).text
        all_movie_info = soup.find_all('div', {'data-test-id': 'encyclopedic-table'})[0].contents
        # all_movie_info = soup.find_all('div', {'class': 'movie-info__table-container'})  # .find('a').text
        for info in all_movie_info:
            """Получаем основные данные о фильме
            """
            # items = info.find('table', {'class': 'info'}).find_all('tr')  # Берем только таблицу INFO о фильме
            # for item in items:
            # i = item.find('td')
            if info.next_element.text.lower() == 'год производства':
                year = info.next_element.next_sibling.next_element.text
                results['year'] = year.strip()
            if info.next_element.text.lower() == 'страна':
                country = info.next_element.next_sibling.text
                results['country'] = country.strip()
            if info.next_element.text.lower() == 'режиссер':
                director = info.next_element.next_sibling.text
                results['director'] = director.strip()
            if info.next_element.text.lower() == 'жанр':
                genre = info.next_element.next_sibling.text.split(',')[0]
                # genre = info.next_element.next_sibling.text
                results['genre'] = genre.strip()
            if info.next_element.text.lower() == 'время':
                    ftime = info.next_element.next_sibling.text
                    results['time'] = ftime.strip()
            """ Получаем актеров фильма
            """
        # items = soup.find('div', {'id': 'actorList'}).find('ul').find_all('a')
        items = soup.find('div', {'class': 'styles_actors__2zt1j'}).find('ul').find_all('a')

        # items = info.find('div', {'id': 'actorList'}).find('ul').find_all('a')
        for i, item in enumerate(items):
            actors[i] = item.text

        try:
            actors.popitem()  # Удаляем актера "...". Во всех фильмах многоточие в конце списка
        except KeyError:
            print(results['film_name'], ': Список актеров пуст!')
        results['actors'] = actors

        # Получаем оценки фильма
        try:
            kinopoisk = soup.find('a', {'class': "film-rating-value"}).text
            imdb = str(soup.find('div', {'class': "film-sub-rating"}).contents[0].contents[-1])
            # imdb = soup.find('div', {'class': "styles_ratingContainer__24Wyy"}).next_element.next_element.nextSibling.\
            #     next_element.text
            kinopoisk = kinopoisk.strip().split('\n')[0]
            imdb = imdb.strip().split()[0]
            results['kinopoisk'] = kinopoisk
            results['imdb'] = imdb
        except AttributeError:
            print('Проблема с оценками у фильма. Возможно их еще нет')
            results.setdefault('kinopoisk', '0')
            results.setdefault('imdb', '0')
        results['description'] = KinopoiskParser.get_description(soup)
        return results

    @staticmethod
    def get_info_hd(content):
        """Получение данных фильма новой версии Кинопоиск HD с помощью BeautifulSoup.

        :param content: объект, который передается в BeautifulSoup
        :return: results - словарь с данными
        """

        soup = BeautifulSoup(content, features="html.parser")
        results = {}
        actors = {}
        # results['film_name'] = soup.find('div',
        #                                  {'class': 'film-header-group film-basic-info__title'}).next_element.next.text
        # results['film_name'] = soup.find('div', {'class': 'film-basic-info__title'}).next_element.next.text
        results['film_name'] = soup.find('span', {'data-tid': '35f45dae'}).text
        # all_movie_info = soup.find('div', {'class': 'film-info-table'})  # film-info-table_color_scheme_grey'})
        all_movie_info = soup.find('div', {'data-test-id': 'encyclopedic-table'})
        # all_movie_info = soup.find('div', {'data-tid': 'bdb69791'})
        for info in all_movie_info.contents:
            # Получаем основные данные о фильме
            if info.contents[0].text == 'Год производства':
                year = re.split(' ', info.contents[1].text)[0]  # Отрезаем, т.к. строка может быть 'год (3 сезона)'
                results['year'] = year.strip()
            if info.contents[0].text == 'Страна':
                country = info.contents[1].text
                results['country'] = country.strip()
            if info.contents[0].text == 'Режиссер':
                director = info.contents[1].text
                results['director'] = director.strip()
            if info.contents[0].text == 'Жанр':
                genre = re.split(',', info.contents[1].text)
                genre = genre[1] if genre[0] == 'ужасы' else genre[0]  # меняем ужасы на следующий жанр
                results['genre'] = genre.strip()
            if info.contents[0].text == 'Время':
                ftime = info.contents[1].text
                results['time'] = ftime.strip()
        # Получаем актеров фильма
        # items = soup.find('div', {'class': "film-crew-block film-basic-info__film-crew"}).contents[0].contents[1].contents[0].contents
        items = soup.find('ul', {'class': "styles_list__I97eu"}).contents
        # items = soup.find('div', {'class': "film-crew-block film-basic-info__film-crew"}).contents[0].contents[1].contents
        for i, item in enumerate(items):
            actors[i] = item.text
        # try:
        #     actors.popitem()  # Удаляем актера "...". Во всех фильмах многоточие в конце списка
        # except KeyError:
        #     print(results['film_name'], ': Список актеров пуст!')
        results['actors'] = actors

        # Получаем оценки фильма
        try:
            kinopoisk = soup.find('a', {'class': "film-rating-value"}).text
            imdb = str(soup.find('div', {'class': "film-sub-rating"}).contents[0].contents[-1])
            results['kinopoisk'] = kinopoisk.strip()
            results['imdb'] = imdb.strip()
        except AttributeError:
            print('Оценок у фильма еще нет')
        results['description'] = KinopoiskParser.get_description(soup)
        return results

    @staticmethod
    def get_description(soup=None):
        """Получаем описание фильма/сериала из готового soup"""
        if not soup:
            print('Не передана soup-страничка для парсинга описания')
            return None
        description = ''
        try:
            items = soup.find('div', {'class', 'styles_filmSynopsis__zLClu'}).find_all('p')
            for i in items:
                description += ' ' + i.contents[0]
        except Exception as e:
            print(e, 'Не удалось получить описание', sep='\n')
        return description

    def get_icon(self):
        id_film = self.result['id_kinopoisk']
        url = f'https://www.kinopoisk.ru/images/sm_film/{id_film}.jpg'

        path_to_icons = ''
        for p in ICON_PATHS:
            if os.path.exists(p):
                path_to_icons = p
                break
        if not path_to_icons:
            print('Не удалось найти путь к папке с иконками')
            return
        req = requests.get(url)
        with open(path_to_icons + f'{id_film}_icon.jpg', 'wb') as icon_file:
            icon_file.write(req.content)
        self.copy_icon_to_serv(path_to_icons, id_film)
        print(f'Загружена иконка для фильма {id_film}')
        return f'main/icons/{id_film}_icon.jpg'

    @staticmethod
    def copy_icon_to_serv(path_to_icons, id_film):
        conn_file = get_dbauth_file_path()
        with open(conn_file) as file:
            db_name = file.readline().rstrip()
            db_user = file.readline().rstrip()
            db_password = file.readline().rstrip()
            db_host = file.readline().rstrip()
        if path_to_icons == ICON_PATHS[1]:
            os.system(f'C:\\Install\\putty\\PSCP.EXE -P 1919 -pw {db_password} {path_to_icons}{id_film}_icon.jpg {db_user}@{db_host}:{ICON_PATHS[2]}{id_film}_icon.jpg')
        elif path_to_icons == ICON_PATHS[0]:
            pass

    def find_film_id(self):
        """Нахождение ID фильма по названию и году на Кинопоиске.

        Получаем страницу поиска по названию self.name_film и ищем подходящий фильм по self.year.
        Записываем его ID в self.id_film.
        Пример строки поиска на Кинопоиске https://www.kinopoisk.ru/index.php?kp_query=мстители

        :return: bool -- удалось ли найти ID фильма
        """
        url = f"https://www.kinopoisk.ru/index.php?kp_query={self.name_film}"
        # get = requests.get(url)  # Request банит

        # Запросы через Selenium
        browser = webdriver.Firefox(
            firefox_profile=r'C:\Users\video\AppData\Roaming\Mozilla\Firefox\Profiles\h3qugs8n.Kinopisk')
        browser.get(url)
        wait = WebDriverWait(browser, 10)
        time.sleep(10)
        get = browser.page_source
        browser.close()
        # content = get.content
        soup = BeautifulSoup(get, features="html.parser")
        results = soup.find_all('p', {'class': 'name'})
        for result in results:
            try:
                year = result.find('span', {'class': 'year'}).text
            except Exception:
                break
            if '-' in year:
                continue
            name = result.find('a').text
            name = name.replace(':', '.')
            if (self.name_film == name.lower()) and (self.year == year):
                self.id_film = result.find('a')['data-id']
                return True
        if not self.id_film:
            print('Фильм не найден')
            return False

    def get_from_file(self, url=None):
        """Получение soup из сохраненной локально страницы фильма, а не из запроса.

        :param url: путь с именем к локально сохраненой странице фильма.
        :return: self.result - словарь в данными фильма
        """
        # url = 'C:/Users/Kenobi/Desktop/Avengers.Endgame.html'
        if not url:
            url = input('Введите путь к файлу html: ')
        with open(url, encoding="utf8") as file:
            text = file.read()
        self.result = self.get_info(text)
        return self.result

    def get_from_kinopoisk_with_id(self):
        """Получение soup из запроса к сайту Кинопоиска по ID фильма.

        Должен быть уже известен self.id_film.
        Ищем его через find_film_id() по названию.

        :return: self.result - словарь в данными фильма
        """
        url = f'https://www.kinopoisk.ru/film/{self.id_film}'
        self.browser.get(url)
        WebDriverWait(self.browser, 10)
        time.sleep(10)
        get = self.browser.page_source
        # content = get.content  # .decode(get.encoding)
        # Проверка на капчу
        '''if 'captcha' in content:
            raise ValueError('Kinopoisk block this IP. Too many requests')'''
        # В случае, если откроется страница KinopoiskHD
        try:
            self.result = self.get_info(get)
        except AttributeError:
            self.result = self.get_info_hd(get)
        # в названиях на Кинопоиске попадаются символ неразрывного пробела, который дает False в сравнении имен
        self.result['film_name'] = self.result['film_name'].replace(chr(160), chr(32))
        self.result['id_kinopoisk'] = self.id_film
        # получаем даты сезонов у сериала
        try:
            # element = self.browser.find_element_by_class_name('table-col-years__seasons')
            element = self.browser.find_elements_by_class_name('styles_years__3VWqc')[0]
            # element = self.browser.find_elements_by_class_name('mC3xgnbJn9r8KpG71O4nw')[1]
            self.browser.execute_script("arguments[0].click();", element)
            time.sleep(10)
            get = self.browser.page_source
            soup = BeautifulSoup(get, features="html.parser")
            years = []
            items = soup.find('div', {'class', 'shadow'}).find_all('a', {'class', 'all'})
            for i in items:
                if '#y' in i['href']:
                    years += i.contents
            self.result["year"] = years
        except Exception:
            print('Даты сезонов не получилось достать.')
        return self.result

    """Методы для поиска через Selenium"""

    def open_selenium(self):
        """Запуск Selenium-браузера(Firefox) и переход на сайт Кинопоиска."""
        url = 'https://www.kinopoisk.ru/'
        # Если нужен профиль в Firefox, то добавить следующий параметр в webdriver.Firefox() (поменять на свой профиль)
        # firefox_profile=r'C:\Users\video\AppData\Roaming\Mozilla\Firefox\Profiles\h3qugs8n.Kinopisk')

        if os.name == 'nt':  # Windows OS
            self.browser = webdriver.Firefox()
        elif os.name == 'posix':  # Mac OS
            driver_path = str(Path(__file__).parent.joinpath('web-drivers').joinpath('chromedriver'))
            self.browser = webdriver.Chrome(driver_path)

        self.browser.get(url)

    def correcting_names(self, name: str):
        """Корректировка названий из базы и сайта для однородного вида.

        :param name: название фильма с сайта
        :return: name
        """
        self.name_film = str(self.name_film).lower()
        self.year = str(self.year)
        name = name.lower()
        # В xlsx-таблице во всех названиях ':' заменено на '.', т.к. названия файлов в Windows не позволяют ставить ':'
        name = name.replace(':', '.')
        # В названиях на Кинопоиске иногда попадаются специальные символы многоточия,
        # которые дают False при сравнении имен. Заменяем их на три точки, как в таблице
        name = name.replace(chr(8230), '...')
        self.name_film = self.name_film.replace(chr(8230), '...')
        # В названиях на Кинопоиске иногда попадаются символ неразрывного пробела, который дает False при сравнении имен
        name = name.replace(chr(160), chr(32))
        # удаляем лишние пробелы в начале и конце
        self.name_film = self.name_film.strip()
        name = name.strip()
        return name

    def find_on_kinopoisk(self, name_film, year_film):
        """Поиск фильма на сайте и возрат его данных.

        В строку поиска на сайте Кинопоиска в Selenium браузере передается название фильма и производится поиск
        нужного фильма на первой странице поисковой выдачи по совпадению имени и года.
        После чего передается код страницы в get_info(), которая возращает нужные данные.

        :param name_film: название фильма из таблицы Films.xlsx
        :param year_film: год фильма из таблицы Films.xlsx
        :return: data: словарь с необходимыми данными о фильме
        """
        self.name_film = name_film
        self.year = year_film
        # запуск поиска на сайте
        search_form = self.browser.find_element_by_name('kp_query')
        search_form.clear()
        search_form.send_keys(self.name_film)
        search_form.submit()
        time.sleep(6)
        try:
            WebDriverWait(self.browser, 10).until(ec.presence_of_all_elements_located((By.CLASS_NAME, 'name')))
        except Exception:
            pass
        # получение и обход результатов поиска
        results = self.browser.find_elements_by_class_name('name')
        for result in results:
            try:
                year = result.find_element_by_class_name('year').text
            except Exception:
                year = ''
                continue
            if '-' in year:  # '-' в сериалах
                continue
            # находим название фильма и отправляем его на обработку
            name = self.correcting_names(result.find_element_by_tag_name('a').text)
            # если названия и годы совпадают, то переходим на страничку фильма
            if (self.name_film == name) and (self.year == year):
                result.find_element_by_tag_name('a').click()
                break
        try:
            # WebDriverWait(self.browser, 10).until(ec.presence_of_element_located((By.CLASS_NAME, 'info')))
            # WebDriverWait(self.browser, 10).until(ec.presence_of_element_located((By.ID, 'actorList')))
            time.sleep(15)
            get = self.browser.page_source
            self.id_film = os.path.basename(os.path.dirname(self.browser.current_url))
            data = KinopoiskParser.get_info(get)
            # в названиях на Кинопоиске попадаются символ неразрывного пробела, который дает False в сравнении имен
            data['film_name'] = data['film_name'].replace(chr(160), chr(32))
            data['id_kinopoisk'] = self.id_film
            return data
        except selenium.common.exceptions.TimeoutException:
            print('Что-то пошло не так с этим фильмом: ', self.name_film)

    def write_data(self):
        """Запись в Films.xlsx данных для последнего фильма.

        Замечен спецэффект, когда max_row больше количества фактически заполненных строк.
        В таком случае необходимо в xlsx вручную удалить пустые строки.

        :return:
        """
        wb = openpyxl.load_workbook(filename=data_file)
        ws = wb.active
        font = Font(name='Times New Roman',
                    size=12,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        # Для фильма и сериала разные алгоритмы поска данных.
        # Фильм ищется по имени и году, а для сериала сразу по ID из таблицы осуществляется переход на страницу сериала
        if ws.cell(row=ws.max_row, column=self._find_column(ws, 'Type')).value == 'Фильм':
            self.result = self.find_on_kinopoisk(name_film=ws.cell(row=ws.max_row,
                                                                   column=self._find_column(ws, 'Name')).value,
                                                 year_film=ws.cell(row=ws.max_row,
                                                                   column=self._find_column(ws, 'Year')).value)
        elif ws.cell(row=ws.max_row, column=self._find_column(ws, 'Type')).value == 'Сериал':
            self.id_film = ws.cell(row=ws.max_row, column=self._find_column(ws, 'Kinopoisk ID')).value
            self.result = self.get_from_kinopoisk_with_id()
            season = ws.cell(row=ws.max_row, column=self._find_column(ws, 'Season')).value
            season = int(re.split(' ', season)[1])
            self.result["year"] = self.result["year"][season - 1]
        # Заполняем таблицу полученными данными
        # Колонки таблицы могут менять свой порядок, поэтому их поиск осуществляется по имени столбцов
        try:
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Year')).value = self.result["year"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Year')).font = font
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Genre')).value = self.result["genre"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Genre')).font = font
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Kinopoisk ID')).value = self.result["id_kinopoisk"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Kinopoisk ID')).font = font
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'IMDB')).value = self.result["imdb"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'IMDB')).font = font
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Kinopoisk')).value = self.result["kinopoisk"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Kinopoisk')).font = font
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Country')).value = self.result["country"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Country')).font = font
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Time')).value = self.result["time"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Time')).font = font
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Director')).value = self.result["director"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Director')).font = font
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Description')).value = self.result["description"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Description')).font = font
            # Актеры - словарь в словаре. Превращаем в единую строку для записи в файл
            actors = ''
            for key, value in self.result['actors'].items():
                actors += value + ", "
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Actors')).value = actors[:-2]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Actors')).font = font
            self.result['icon'] = self.get_icon()
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Icon')).value = self.result["icon"]
            ws.cell(row=ws.max_row, column=self._find_column(ws, 'Icon')).font = font
        except KeyError as e:
            print('Нет значения: ', e)
        self.name_film = ws.cell(row=ws.max_row, column=self._find_column(ws, 'Name')).value
        wb.save(data_file)
        wb.close()
        print(f'Получены и записаны данные с Кинопоиска для {self.name_film}')

    @staticmethod
    def _find_column(sheet, name):
        """
        Поиск номера столбца по имени.

        :param sheet: рабочий лист в xlsx
        :param name: имя столбца для поиска его номера
        :return: номер столбца
        """
        for cell in sheet[1]:
            if cell.value == name:
                return cell.col_idx

    def close_selenium(self):
        """Закрытие Selenium браузера"""
        self.browser.close()


if __name__ == '__main__':
    identificator = input('Введите ID фильма: ')
    film = KinopoiskParser(identificator)
    # results = film.get_from_file()
    output = film.get_from_kinopoisk_with_id()
    print(output)
