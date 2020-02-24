import os
import shutil
import psutil
import openpyxl
import xlrd

arc_disk = input('Введите букву диска с архивом: ').upper() + ':'
chief_disk = input('Введите букву диска для записи шефу: ').upper() + ':'
serv_disk = input('Введите букву диска для записи на видеосервер: ').upper() + ':'

wname_serial = ''


def check_subtitr_film(name, extension, a=arc_disk + '/New/'):
    """ Проверка на существование дубликата без субтитров и посторонних дорожек
    """
    if (os.path.exists(a + name + '(1)' + extension)) or (os.path.exists(a + name + ' (1)' + extension)):
        return True
    else:
        return False


def write_films_xlsx(name, season=None, jenre=None, year=None):
    """Запись в файл Films.xlsx Сериал-Сезон или Фильм-Жанр-Год
       write_films_xlsx(name_name, s, spl[0], spl[1])
    """
    wb = openpyxl.load_workbook(filename='C:/install/Films.xlsx')
    ws = wb.active
    rb = xlrd.open_workbook('C:\install\Films.xlsx')
    sheet = rb.sheet_by_index(0)
    ws["A" + str(sheet.nrows + 1)] = name
    if season: ws["B" + str(sheet.nrows + 1)] = 'Сезон ' + str(season)
    if jenre: ws["B" + str(sheet.nrows + 1)] = jenre
    if year: ws["C" + str(sheet.nrows + 1)] = year
    wb.save('C:\install\Films.xlsx')
    return True


def newname_for_copy(name):
    """ Переименование фильмов, оканчивающихся на (1). name  = os.path.splitext(spl[2]) = [название, расширение]
    """
    new = "".join(name)
    if name[0].endswith('(1)'):  # Если это копия без субтитров и посторонних дорожек, то обрезаем в названии (1)
        new = name[0][:-4].rstrip() + name[1]
    return new


def name_and_season(n, disk, addr, file):
    season = 1
    if n.endswith('сезон)'):
        season = n[n.find('(') + 1:-6].strip()
        if not season.isdigit():
            season = season[0]
            season = int(season)
        name = n[:n.find('(')].strip()
    else:
        name = n

    while os.path.exists(disk + '/Сериалы/' + name + '/Season ' + str(season)):
        if ( file != [] and not os.path.exists(disk + '/Сериалы/' + name + '/Season ' + str(season) + '/Episode 1' +
                os.path.splitext(file[0])[1])) or (file != [] and os.path.getsize(addr + '\\' + file[0]) != os.path.getsize(
                disk + '/Сериалы/' + name + '/Season ' + str(season) + '/Episode 1' +
                os.path.splitext(file[0])[1])):
            season += 1
        else:
            break
    return name, season


"""Запись на диск для видеосервера
"""

work_path = (arc_disk + '/Convert', arc_disk + '/New')
for work_path_i in work_path:
    if os.path.exists(work_path_i):
        os.chdir(work_path_i)
    else:
        print(work_path_i + ' не существует')
        continue
    s = 1  # номер сезона
    wname_serial = ""
    for adress, dirs, files in os.walk(os.getcwd()):  # получаем (текущая директория, подкаталоги, файлы)
        s = 1  # номер сезона
        if adress == os.getcwd():  # проверка фильм или сериал(фильмы лежат в корне, сериалы в поддиректориях)
            for f in files:
                new_name = ''
                file_size = os.path.getsize(adress + '\\' + f) / 1024 / 1024 / 1024  # размер файла
                free = psutil.disk_usage(serv_disk).free / (1024 * 1024 * 1024)  # свободное место на диске
                spl = f.split('_')  # Разбиваем длинное название на [жанр, год, имя]
                if (file_size < free) and (file_size < 9.0) and not check_subtitr_film(*os.path.splitext(spl[2])):
                    # если фильм помещается на диск и его размер меньше 9 ГБ
                    spl[0] = spl[0].capitalize()  # Первый символ строки большой, остальные маленькие
                    if os.path.exists(serv_disk + '/Фильмы/' + spl[0] + '/' + spl[2]):
                        continue  # если файл уже есть на диске, то пропускаем его /не отрабатывает для (1)
                    if os.path.exists(serv_disk + '/Фильмы/' + spl[0]) is False:
                        os.makedirs(serv_disk + '/Фильмы/' + spl[0])  # создаем директорию Жанр к будущему файлу
                    new_name = newname_for_copy(os.path.splitext(spl[2]))  # обрезаем имя для (1)
                    if os.path.exists(serv_disk + '/Фильмы/' + spl[0] + '/' + new_name):
                        continue   # Если файл уже записан, то пропускаем
                    else:
                        z = shutil.copy(f, serv_disk + '/Фильмы/' + spl[0] + '/' + new_name)
                    print(adress + '\\' + f + ' СКОПИРОВАНО В ' + z)
                    new_name = os.path.splitext(new_name)[0]  # name.ext => name

                    film_jenre = open(serv_disk + '/Фильмы/' + spl[0] + '/' + spl[0] + '.doc', 'a+')
                    film_jenre.write(new_name + '\t' + spl[1] + '\n')  # запись в файл жанров

                    write_films_xlsx(name=new_name, jenre=spl[0], year=spl[1])
        else:
            files.sort(key=len)
            if os.path.split(adress)[0] == os.getcwd():
                wname_serial = adress.split('\\')[-1].capitalize()  # отрезаем название сериала из пути
            name_serial, s = name_and_season(wname_serial, serv_disk, adress, files)
            """ Запись сериала в Films.xlsx, только если находимся в папке "номер сезона" или "название сериала", 
                но нет папки сезона. И если это запись, а не повторный проход.
            """
            if (os.path.split(adress)[0] != os.getcwd() or not dirs) and (
                    not os.path.exists(serv_disk + '/Сериалы/' + name_serial + '/Season ' + str(s))):
                write_films_xlsx(name=name_serial, season=s)
            e = 1  # номер эпизода
            for f in files:  # создаем директории к будущему файлу
                if os.path.exists(serv_disk + '/Сериалы/' + '/' + name_serial + '/Season ' + str(s)) is False:
                    os.makedirs(serv_disk + '/Сериалы/' + '/' + name_serial + '/Season ' + str(s))
                file_size = os.path.getsize(adress) / 1024 / 1024 / 1024  # размер файла
                free = psutil.disk_usage(serv_disk).free / (1024 * 1024 * 1024)
                if (file_size < free) and (file_size < 9.0):
                    if os.path.exists(
                            serv_disk + '/Сериалы/' + name_serial + '/Season ' + str(s) + '/' + 'Episode ' + str(e) +
                            os.path.splitext(f)[1]):
                        e += 1
                        if e == 13:
                            e += 1
                        continue
                    z = shutil.copy(adress + '/' + f, serv_disk + '/Сериалы/' + name_serial + '/Season ' + str(
                        s) + '/' + 'Episode ' + str(e) + os.path.splitext(f)[1])
                    e += 1
                    if e == 13:
                        e += 1
                    print(adress + '\\' + f + ' СКОПИРОВАНО В ' + z)

""" Запись на диск шефу
"""

os.chdir(arc_disk + '/New')

for adress, dirs, files in os.walk(os.getcwd()):  # получаем (текущая директория, подкаталоги, файлы)
    s = 1  # номер сезона
    if adress == os.getcwd():  # проверка фильм или сериал(фильмы лежат в корне, сериалы в поддиректориях)
        for f in files:
            new_name = ''
            file_size = os.path.getsize(adress + '\\' + f) / 1024 / 1024 / 1024  # размер файла
            free = psutil.disk_usage(chief_disk).free / (1024 * 1024 * 1024)  # свободное место на диске
            spl = f.split('_')  # Разбиваем длинное название на [жанр, год, имя]
            if (file_size < free) and not check_subtitr_film(*os.path.splitext(spl[2])):
                # если фильм помещается на диск и нет его копий без субтитров
                spl[0] = spl[0].capitalize()  # Первый символ строки большой, остальные маленькие
                if os.path.exists(chief_disk + '/' + spl[2]):  # если файл уже есть на диске, то пропускаем его
                    continue
                if (os.path.splitext(spl[2])[0][-1] == ')') and (os.path.splitext(spl[2])[0][-2] == '1') and (
                        os.path.splitext(spl[2])[0][-3] == '('):
                    for i in range(0, len(os.path.splitext(spl[2])[0]) - 4):
                        new_name = new_name + os.path.splitext(spl[2])[0][i]
                    if os.path.exists(chief_disk + '/' + new_name + os.path.splitext(spl[2])[1]):
                        continue
                    z = shutil.copy(f, chief_disk + '/' + new_name + os.path.splitext(spl[2])[1])
                else:
                    z = shutil.copy(f, chief_disk + '/' + spl[2])
                print(adress + '\\' + f + ' СКОПИРОВАНО В ' + z)
    else:
        files.sort(key=len)
        if os.path.split(adress)[0] == os.getcwd():
            wname_serial = adress.split('\\')[-1].capitalize()  # отрезаем название сериала из пути
        name_serial, s = name_and_season(wname_serial, chief_disk, adress, files)
        e = 1  # номер эпизода
        for f in files:
            if os.path.exists(chief_disk + '/' + name_serial + '/Season ' + str(s)) is False:
                os.makedirs(chief_disk + '/' + name_serial + '/Season ' + str(s))  # создаем директории к будущему файлу
            file_size = os.path.getsize(adress) / 1024 / 1024 / 1024  # размер файла
            free = psutil.disk_usage(chief_disk).free / (1024 * 1024 * 1024)
            if file_size < free:
                g = os.path.exists(chief_disk + '/' + name_serial + '/Season ' + str(s) + '/' + 'Episode ' + str(e))
                if os.path.exists(chief_disk + '/' + name_serial + '/Season ' + str(s) + '/' + 'Episode ' + str(e) +
                                  os.path.splitext(f)[1]):
                    e += 1
                    if e == 13:
                        e += 1
                    continue
                z = shutil.copy(adress + '/' + f,
                                chief_disk + '/' + name_serial + '/Season ' + str(s) + '/' + 'Episode ' + str(e) +
                                os.path.splitext(f)[1])
                e += 1
                if e == 13:
                    e += 1
                print(adress + '\\' + f + ' СКОПИРОВАНО В ' + z)
    # if dirs == []: s += 1
